"""
ComReg Tower Database Builder

Downloads and processes official Irish cell tower data from ComReg.
Converts ITM coordinates to WGS84 GPS and builds SQLite database.

Data Source: https://www.comreg.ie/industry/radio-spectrum/licensing/
License: Creative Commons Attribution 4.0
"""

import os
import sqlite3
from pathlib import Path
import openpyxl
from pyproj import Transformer
import logging

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Coordinate transformer: ITM (EPSG:2157) → WGS84 (EPSG:4326)
transformer = Transformer.from_crs("EPSG:2157", "EPSG:4326", always_xy=True)

def convert_itm_to_wgs84(easting, northing):
    """
    Convert Irish Transverse Mercator (ITM) coordinates to WGS84 lat/lon.
    
    Args:
        easting: ITM Easting coordinate (meters)
        northing: ITM Northing coordinate (meters)
    
    Returns:
        tuple: (latitude, longitude) in WGS84 decimal degrees
    """
    try:
        lon, lat = transformer.transform(easting, northing)
        return lat, lon
    except Exception as e:
        logger.error(f"Coordinate conversion failed for ({easting}, {northing}): {e}")
        return None, None


def parse_excel_file(file_path, operator_name):
    """
    Parse ComReg Excel file and extract tower data.
    
    Args:
        file_path: Path to Excel file
        operator_name: Operator name (Vodafone, Three, Eir, etc.)
    
    Returns:
        list: Tower records with converted coordinates
    """
    logger.info(f"📄 Parsing {operator_name} file: {file_path}")
    
    if not os.path.exists(file_path):
        logger.warning(f"⚠️  File not found: {file_path}")
        return []
    
    towers = []
    
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook.active
        
        # Find header row (usually row 1, but may vary)
        headers = None
        header_row = 1
        
        for row_idx in range(1, 10):  # Check first 10 rows for headers
            row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            if row and any(h and 'easting' in str(h).lower() for h in row):
                headers = [str(h).strip().lower() if h else '' for h in row]
                header_row = row_idx
                break
        
        if not headers:
            logger.error(f"❌ Could not find header row in {file_path}")
            return []
        
        logger.info(f"📊 Headers: {headers[:5]}...")
        
        # Find column indices
        easting_col = next((i for i, h in enumerate(headers) if 'easting' in h), None)
        northing_col = next((i for i, h in enumerate(headers) if 'northing' in h), None)
        site_col = next((i for i, h in enumerate(headers) if 'site' in h or 'identity' in h), None)
        services_col = next((i for i, h in enumerate(headers) if 'service' in h), None)
        
        if easting_col is None or northing_col is None:
            logger.error(f"❌ Could not find Easting/Northing columns in {file_path}")
            return []
        
        # Parse data rows
        row_count = 0
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not row[easting_col] or not row[northing_col]:
                continue
            
            try:
                easting = float(row[easting_col])
                northing = float(row[northing_col])
                site_id = str(row[site_col]) if site_col is not None and row[site_col] else f"{operator_name}_{row_count}"
                services = str(row[services_col]) if services_col is not None and row[services_col] else "Unknown"
                
                # Convert coordinates
                lat, lon = convert_itm_to_wgs84(easting, northing)
                
                if lat and lon:
                    towers.append({
                        'site_id': site_id,
                        'operator': operator_name,
                        'latitude': lat,
                        'longitude': lon,
                        'easting': easting,
                        'northing': northing,
                        'services': services
                    })
                    row_count += 1
                    
            except (ValueError, TypeError) as e:
                continue  # Skip invalid rows
        
        logger.info(f"✅ Parsed {row_count} towers from {operator_name}")
        workbook.close()
        return towers
        
    except Exception as e:
        logger.error(f"❌ Error parsing {file_path}: {e}")
        return []


def build_database(towers, db_path):
    """
    Build SQLite database from tower data.
    
    Args:
        towers: List of tower records
        db_path: Path to SQLite database file
    """
    logger.info(f"🗄️  Building database: {db_path}")
    
    # Create database
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Create table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS towers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            site_id TEXT,
            operator TEXT,
            latitude REAL,
            longitude REAL,
            easting REAL,
            northing REAL,
            services TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Create spatial index for faster queries
    cursor.execute('''
        CREATE INDEX IF NOT EXISTS idx_location ON towers (latitude, longitude)
    ''')
    
    cursor.execute('''
        CREATE INDEX IF NOT EXISTS idx_operator ON towers (operator)
    ''')
    
    # Insert towers
    cursor.executemany('''
        INSERT INTO towers (site_id, operator, latitude, longitude, easting, northing, services)
        VALUES (:site_id, :operator, :latitude, :longitude, :easting, :northing, :services)
    ''', towers)
    
    conn.commit()
    
    # Get stats
    cursor.execute('SELECT COUNT(*) FROM towers')
    total = cursor.fetchone()[0]
    
    cursor.execute('SELECT operator, COUNT(*) FROM towers GROUP BY operator')
    stats = cursor.fetchall()
    
    logger.info(f"✅ Database created with {total} towers:")
    for operator, count in stats:
        logger.info(f"   - {operator}: {count} sites")
    
    conn.close()


def main():
    """Main function to build ComReg tower database."""
    
    print("\n" + "="*60)
    print("  📡 ComReg Tower Database Builder")
    print("="*60 + "\n")
    
    # Setup paths
    project_root = Path(__file__).parent.parent
    data_dir = project_root / "data" / "comreg"
    data_dir.mkdir(parents=True, exist_ok=True)
    
    db_path = project_root / "data" / "comreg_towers.db"
    
    print("📂 Please download ComReg Excel files manually:")
    print("   https://www.comreg.ie/industry/radio-spectrum/licensing/")
    print("   → Mobile Licences")
    print("   → Download Excel files for each operator\n")
    print(f"💾 Save them to: {data_dir}\n")
    print("📝 Example filenames:")
    print("   - vodafone_800mhz.xlsx")
    print("   - three_2100mhz.xlsx")
    print("   - eir_900mhz.xlsx")
    print("\n" + "="*60 + "\n")
    
    # Scan for Excel files in data directory
    excel_files = list(data_dir.glob("*.xlsx")) + list(data_dir.glob("*.xls"))
    
    if not excel_files:
        print("⚠️  No Excel files found in data/comreg/")
        print("\nPlease download the files and run this script again.")
        return
    
    print(f"📁 Found {len(excel_files)} Excel file(s):\n")
    for f in excel_files:
        print(f"   - {f.name}")
    print()
    
    # Parse all files
    all_towers = []
    
    for file_path in excel_files:
        # Try to guess operator from filename
        filename_lower = file_path.stem.lower()
        if 'vodafone' in filename_lower:
            operator = 'Vodafone Ireland'
        elif 'three' in filename_lower:
            operator = 'Three Ireland'
        elif 'eir' in filename_lower:
            operator = 'Eir'
        else:
            operator = file_path.stem  # Use filename as operator name
        
        towers = parse_excel_file(str(file_path), operator)
        all_towers.extend(towers)
    
    if not all_towers:
        print("❌ No tower data extracted. Check Excel file format.")
        return
    
    # Build database
    build_database(all_towers, str(db_path))
    
    print(f"\n✅ Database created: {db_path}")
    print(f"📊 Total towers: {len(all_towers)}")
    print("\n" + "="*60 + "\n")


if __name__ == "__main__":
    main()
